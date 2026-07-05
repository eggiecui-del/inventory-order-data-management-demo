from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from database import DATABASE_URL, get_connection, get_product_id_by_code, init_db


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "sample_inventory_data.xlsx"


def normalize_text(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def normalize_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def normalize_float(value, default=0.0):
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_datetime(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = normalize_text(value)
    if len(text) == 10:
        return f"{text} 00:00:00"
    return text or datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def supplier_code_for_name(name):
    cleaned = "".join(char for char in name.upper() if char.isalnum())
    return f"SUP-{cleaned[:24] or 'UNKNOWN'}"


def upsert_supplier(conn, supplier_name, supplier_city):
    if not supplier_name:
        return None
    supplier_code = supplier_code_for_name(supplier_name)
    row = conn.execute(
        """
        INSERT INTO suppliers (supplier_code, supplier_name, city)
        VALUES (?, ?, ?)
        ON CONFLICT (supplier_code) DO UPDATE SET
            supplier_name = EXCLUDED.supplier_name,
            city = EXCLUDED.city
        RETURNING id
        """,
        (supplier_code, supplier_name, supplier_city),
    ).fetchone()
    return row["id"]


def read_sheet_rows(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(value) for value in rows[0]]
    return [
        {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]


def import_workbook(excel_path=EXCEL_PATH, database_url=DATABASE_URL, reset=True):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    required_sheets = {"product_info", "inventory", "inventory_log"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"Missing sheets: {', '.join(sorted(missing_sheets))}")

    init_db(database_url)
    summary = {"products": 0, "inventory": 0, "inventory_logs": 0, "skipped": []}

    with get_connection(database_url) as conn:
        if reset:
            conn.execute("DELETE FROM audit_logs")
            conn.execute("DELETE FROM order_items")
            conn.execute("DELETE FROM orders")
            conn.execute("DELETE FROM customers")
            conn.execute("DELETE FROM inventory_logs")
            conn.execute("DELETE FROM inventory")
            conn.execute("DELETE FROM products")
            conn.execute("DELETE FROM suppliers")
            conn.execute("DELETE FROM users")

        seen_codes = set()
        for row in read_sheet_rows(workbook, "product_info"):
            product_code = normalize_text(row.get("product_code"))
            product_name = normalize_text(row.get("product_name"))
            unit = normalize_text(row.get("unit"))
            if not product_code or not product_name or not unit:
                summary["skipped"].append(f"Missing required product fields: {product_code or product_name or 'blank row'}")
                continue
            if product_code in seen_codes or get_product_id_by_code(conn, product_code):
                summary["skipped"].append(f"Duplicate product code skipped: {product_code}")
                continue
            seen_codes.add(product_code)

            created_at = normalize_datetime(row.get("created_at"))
            supplier_name = normalize_text(row.get("supplier_name"))
            supplier_city = normalize_text(row.get("supplier_city"))
            supplier_id = upsert_supplier(conn, supplier_name, supplier_city)
            conn.execute(
                """
                INSERT INTO products (
                    product_code, product_name, category, brand, model, unit,
                    cost_price, sale_price, supplier_id, supplier_name, supplier_city,
                    usage_scene, remark, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product_code,
                    product_name,
                    normalize_text(row.get("category")),
                    normalize_text(row.get("brand")),
                    normalize_text(row.get("model")),
                    unit,
                    normalize_float(row.get("cost_price")),
                    normalize_float(row.get("sale_price")),
                    supplier_id,
                    supplier_name,
                    supplier_city,
                    normalize_text(row.get("usage_scene")),
                    normalize_text(row.get("remark")),
                    created_at,
                    created_at,
                ),
            )
            summary["products"] += 1

        for row in read_sheet_rows(workbook, "inventory"):
            product_code = normalize_text(row.get("product_code"))
            product_id = get_product_id_by_code(conn, product_code)
            if not product_id:
                summary["skipped"].append(f"Inventory row skipped, product not found: {product_code}")
                continue
            source_inventory_code = normalize_text(row.get("inventory_code"))
            remark = normalize_text(row.get("remark"))
            if source_inventory_code:
                remark = f"{remark}; source inventory code: {source_inventory_code}" if remark else f"source inventory code: {source_inventory_code}"
            safety_stock = normalize_int(row.get("safety_stock"))
            minimum_stock = normalize_int(row.get("minimum_stock"), max(0, safety_stock // 2))
            conn.execute(
                """
                INSERT INTO inventory (
                    product_id, current_quantity, location, minimum_stock, safety_stock,
                    last_updated_at, remark
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(product_id) DO UPDATE SET
                    current_quantity = excluded.current_quantity,
                    location = excluded.location,
                    minimum_stock = excluded.minimum_stock,
                    safety_stock = excluded.safety_stock,
                    last_updated_at = excluded.last_updated_at,
                    remark = excluded.remark
                """,
                (
                    product_id,
                    normalize_int(row.get("current_quantity")),
                    normalize_text(row.get("location")),
                    minimum_stock,
                    safety_stock,
                    normalize_datetime(row.get("last_updated_at")),
                    remark,
                ),
            )
            summary["inventory"] += 1

        for row in read_sheet_rows(workbook, "inventory_log"):
            product_code = normalize_text(row.get("product_code"))
            product_id = get_product_id_by_code(conn, product_code)
            if not product_id:
                summary["skipped"].append(f"Inventory log skipped, product not found: {product_code}")
                continue
            change_type = normalize_text(row.get("change_type"))
            if change_type not in {"stock_in", "stock_out", "adjustment"}:
                summary["skipped"].append(f"Invalid change type skipped: {product_code} {change_type}")
                continue
            record_code = normalize_text(row.get("log_code"))
            note = normalize_text(row.get("note"))
            if record_code:
                note = f"{note}; source log code: {record_code}" if note else f"source log code: {record_code}"
            conn.execute(
                """
                INSERT INTO inventory_logs (
                    product_id, change_type, quantity_change, quantity_before,
                    quantity_after, reason, note, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product_id,
                    change_type,
                    normalize_int(row.get("quantity_change")),
                    normalize_int(row.get("quantity_before")),
                    normalize_int(row.get("quantity_after")),
                    normalize_text(row.get("reason")),
                    note,
                    normalize_datetime(row.get("created_at")),
                ),
            )
            summary["inventory_logs"] += 1

        conn.commit()

    return summary


if __name__ == "__main__":
    result = import_workbook()
    print("Excel import finished")
    print(f"products: {result['products']}")
    print(f"inventory: {result['inventory']}")
    print(f"inventory_logs: {result['inventory_logs']}")
    if result["skipped"]:
        print("skipped:")
        for item in result["skipped"]:
            print(f"- {item}")
